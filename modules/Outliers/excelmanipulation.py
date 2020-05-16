import openpyxl as xl

from openpyxl.chart import BarChart,Reference




wb=xl.load_workbook('transactions.xlsx')

sheet=wb['Sheet1']


# cell=sheet['a1']

# cellu=sheet.cell(2,3)

# print(cellu.value)


for row in range(2,sheet.max_row +1):

    cellu=sheet.cell(row,3)
   
    calculated= cellu.value*10
    
    print(cellu.value)

    calculatedcell=sheet.cell(row,4)

    calculatedcell.value=calculated



values=Reference(sheet,
        min_row=2,
        max_row=sheet.max_row,
        min_col=4,
        max_col=4

)

chart= BarChart()

chart.add_data(values)

sheet.add_chart(chart,'E2')


wb.save('transaction1.xlsx')

    

 



